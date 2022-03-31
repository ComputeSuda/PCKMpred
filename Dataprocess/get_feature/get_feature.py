from CrossTalk_class import CrossTalk
from read_Excel import readExcel
from write_Excel import writeExcel
import openpyxl
from openpyxl.styles import Font

# Origin dataset Cross-talk/Mutation
positiveset_file = '../Dataset/Cross-talk/New_dataset_positive.xlsx'
negativeset_file = '../Dataset/Cross-talk/New_dataset_negative.xlsx'
# positiveset_file = r'C:\Users\leishen\Desktop\PTM_Cross_talk_Imbalanced_learning\dataset\ptm cross-talk\PositiveDataset.xlsx'
# negativeset_file = r'C:\Users\leishen\Desktop\PTM_Cross_talk_Imbalanced_learning\dataset\ptm cross-talk\NegativeDataset.xlsx'

# feature filepath
feature_path = '../../Feature_extraction'
# feature_path = r'C:\Users\leishen\Desktop\PTM_Cross_talk_Imbalanced_learning\feature\results_feature'
# Structural features
cij_path = feature_path + r'Bio3d\cij'  # betweenness, clossness, degree, cluster, diversity, eccentricity, strength,eigen_centrality, page_rank
enm_anm_cancer_cc_path = feature_path + r'\enm\results\anm\cc'  # 5_per,5_per_20_per, 20_per_50_per, greater_60_per, top3
enm_anm_cancer_prs_path = feature_path + r'\enm\results\anm\prs'  # effectiveness, prs, sensitivity
enm_anm_cancer_sq_path = feature_path + r'\enm\results\anm\sq'  # sq
enm_anm_cancer_stiffness_path = feature_path + r'\enm\results\anm\stiffness'  # stiffness
enm_gnm_cancer_cc_path = feature_path + r'\enm\results\gnm\cc'  # 5_per,5_per_20_per, 20_per_50_per, greater_60_per, top3
enm_gnm_cancer_eigenvector_path = feature_path + r'\enm\results\gnm\eigenvector'  # eigenvector_20, eigenvector_all, eigenvector_top3
enm_gnm_cancer_prs_path = feature_path + r'\enm\results\gnm\prs'  # effectiveness, prs, sensitivity
enm_gnm_cancer_sq_path = feature_path + r'\enm\results\gnm\sq'  # sq
# Sequence features
# evol_dirinfo = feature_path + r'\results\dirinfo'
evol_entropy = feature_path + r'\results\entropy'
evol_mifc = feature_path + r'\results\mifc'
evol_mifn = feature_path + r'\results\mifn'
evol_mutinfo = feature_path + r'\results\mutinfo'
evol_occupancy = feature_path + r'\results\occupancy'
evol_omes = feature_path + r'\results\omes'
evol_sca = feature_path + r'\results\sca'


def get_feature(file, featurefile):
    sheet, rows, cols = readExcel(file, 'Sheet1')
    print(sheet, rows, cols)  # <xlrd.sheet.Sheet object at 0x0000024792E09EC8> 10722 21

    titles = ['Protein_name', 'UniprotId', 'Ptm1UniprotSite', "Ptm1PDBPosition", 'Ptm1PDBPositionAbsolute',
              "Ptm1Type", "Ptm2UniprotSite", "Ptm2PDBPosition", 'Ptm2PDBPositionAbsolute', 'Ptm2Type', 'PdbChain',
              'sequence_distance', 'structure_distance',
              'betweenness_cij_min', 'betweenness_cij_max', 'betweenness_cij_avg',
              'closeness_cij_min', 'closeness_cij_max', 'closeness_cij_avg',
              'cluster_cij_min', 'cluster_cij_max', 'cluster_cij_avg',
              'degree_cij_min', 'degree_cij_max', 'degree_cij_avg',
              'diversity_cij_min', 'diversity_cij_max', 'diversity_cij_avg',
              'eccentricity_cij_min', 'eccentricity_cij_max', 'eccentricity_cij_avg',
              'eigen_cij_min', 'eigen_cij_max', 'eigen_cij_avg',
              'page_cij_min', 'page_cij_max', 'page_cij_avg',
              'strength_cij_min', 'strength_cij_max', 'strength_cij_avg',
              'anm_cc_5_per', 'anm_cc_5_per_20_per', 'anm_cc_20_per_50_per', 'anm_cc_greater_60_per', 'anm_cc_top3',
              'anm_prs_all',
              'anm_effectiveness_min', 'anm_effectiveness_max', 'anm_effectiveness_avg',
              'anm_sensitivity_min', 'anm_sensitivity_max', 'anm_sensitivity_avg',
              'anm_sq_min', 'anm_sq_max', 'anm_sq_avg',
              'anm_stiffness',
              'gnm_cc_5_per', 'gnm_cc_5_per_20_per', 'gnm_cc_20_per_50_per', 'gnm_cc_greater_60_per', 'gnm_cc_top3',
              'gnm_eigenvectors_20_min', 'gnm_eigenvectors_20_max', 'gnm_eigenvectors_20_avg',
              'gnm_eigenvectors_all_min', 'gnm_eigenvectors_all_max', 'gnm_eigenvectors_all_avg',
              'gnm_eigenvectors_top3_min', 'gnm_eigenvectors_top3_max', 'gnm_eigenvectors_top3_avg',
              'gnm_prs',
              'gnm_effectiveness_min', 'gnm_effectiveness_max', 'gnm_effectiveness_avg',
              'gnm_sensitivity_min', 'gnm_sensitivity_max', 'gnm_sensitivity_avg',
              'gnm_sq_min', 'gnm_sq_max', 'gnm_sq_avg',
              'evol_entropy_min', 'evol_entropy_max', 'evol_entropy_avg',
              'evol_mifc',
              'evol_mifn',
              'evol_mutinfo',
              'evol_occupancy_min', 'evol_occupancy_max', 'evol_occupancy_avg',
              'evol_omes',
              'evol_sca'
              ]

    for row in range(1, rows + 1):
        # for row in range(1, 5):
        if row == 1:
            Ptm_CrossTalk_excel = openpyxl.Workbook()
            sheet1 = Ptm_CrossTalk_excel.create_sheet('Sheet1', 0)
            for i in range(1, len(titles) + 1):
                sheet1.cell(row, i, titles[i - 1])
        else:
            features = []
            print(row)
            ProteinName = str(sheet.row_values(row - 1)[0])
            UniprotId = str(sheet.row_values(row - 1)[1])
            Ptm1UniprotSite = int(sheet.row_values(row - 1)[2][1:])
            if sheet.row_values(row - 1)[3] == 'None':
                Ptm1PdbPosition = str(sheet.row_values(row - 1)[3])
            else:
                Ptm1PdbPosition = int(sheet.row_values(row - 1)[3])
            if sheet.row_values(row - 1)[4] == 'None':
                Ptm1PdbPositionAbsolute = str(sheet.row_values(row - 1)[4])
            else:
                Ptm1PdbPositionAbsolute = int(sheet.row_values(row - 1)[4])
            Ptm1Type = str(sheet.row_values(row - 1)[5])
            Ptm2UniprotSite = int(sheet.row_values(row - 1)[6][1:])
            if sheet.row_values(row - 1)[7] == 'None':
                Ptm2PdbPosition = str(sheet.row_values(row - 1)[7])
            else:
                Ptm2PdbPosition = int(sheet.row_values(row - 1)[7])
            if sheet.row_values(row - 1)[8] == 'None':
                Ptm2PdbPositionAbsolute = str(sheet.row_values(row - 1)[8])
            else:
                Ptm2PdbPositionAbsolute = int(sheet.row_values(row - 1)[8])
            Ptm2Type = str(sheet.row_values(row - 1)[9])
            PdbChain = str(sheet.row_values(row - 1)[10])
            Sequence_distance = int(sheet.row_values(row - 1)[11])
            if sheet.row_values(row - 1)[12] == 'None':
                Structure_distance = str(sheet.row_values(row - 1)[12])
            else:
                Structure_distance = int(sheet.row_values(row - 1)[12])
            Ptm_CrossTalk = CrossTalk(ProteinName, UniprotId, Ptm1UniprotSite, Ptm1PdbPosition,
                                       Ptm1PdbPositionAbsolute, Ptm1Type, Ptm2UniprotSite,
                                       Ptm2PdbPosition, Ptm2PdbPositionAbsolute, Ptm2Type, PdbChain, Sequence_distance,
                                       Structure_distance)
            Ptm_CrossTalk.pdbchain_feature_cij(cij_path)
            Ptm_CrossTalk.pdbchain_feature_anm_cc(enm_anm_cancer_cc_path)
            Ptm_CrossTalk.pdbchain_feature_anm_prs(enm_anm_cancer_prs_path)
            Ptm_CrossTalk.pdbchain_feature_anm_sq(enm_anm_cancer_sq_path)
            Ptm_CrossTalk.pdbchain_feature_anm_stiffness(enm_anm_cancer_stiffness_path)
            Ptm_CrossTalk.pdbchain_feature_gnm_cc(enm_gnm_cancer_cc_path)
            Ptm_CrossTalk.pdbchain_feature_gnm_eigenvector(enm_gnm_cancer_eigenvector_path)
            Ptm_CrossTalk.pdbchain_feature_gnm_prs(enm_gnm_cancer_prs_path)
            Ptm_CrossTalk.pdbchain_feature_gnm_sq(enm_gnm_cancer_sq_path)
            # print(Ptm_mutation.UniprotPosition)
            # Ptm_mutation.uninprot_feature_evol_dirinfo(evol_dirinfo)
            Ptm_CrossTalk.uninprot_feature_evol_entropy(evol_entropy)
            Ptm_CrossTalk.uninprot_feature_evol_mifc(evol_mifc)
            Ptm_CrossTalk.uninprot_feature_evol_mifn(evol_mifn)
            Ptm_CrossTalk.uninprot_feature_evol_mutinfo(evol_mutinfo)
            Ptm_CrossTalk.uninprot_feature_evol_occupancy(evol_occupancy)
            Ptm_CrossTalk.uninprot_feature_evol_omes(evol_omes)
            Ptm_CrossTalk.uninprot_feature_evol_sca(evol_sca)
            # print(Ptm_mutation.items[0])
            # print(Ptm_mutation.uninprot_feature, Ptm_mutation.pdbchain_feature)
            # print(list(Ptm_mutation.pdbchain_feature.keys()) + list(Ptm_mutation.uninprot_feature.keys()))
            features.extend(
                [ProteinName, UniprotId, Ptm1UniprotSite, Ptm1PdbPosition,
                 Ptm1PdbPositionAbsolute, Ptm1Type, Ptm2UniprotSite,
                 Ptm2PdbPosition, Ptm2PdbPositionAbsolute, Ptm2Type, PdbChain, Sequence_distance,
                 Structure_distance])
            # features.extend(Ptm_mutation.items[0][:8])
            # features.extend([Ptm_mutation.MutationPdbPosition, Ptm_mutation.MutationPdbPositionAbsolute,
            #                  Ptm_mutation.MutationUniprotPosition])
            # features.extend(Ptm_mutation.items[0][8:])
            # print(type(Ptm_mutation.items))
            # print(Ptm_mutation.items)
            # features.extend(Ptm_mutation.items[0])
            # print(list(Ptm_mutation.pdbchain_feature.values()))
            # print(list(Ptm_mutation.uninprot_feature.values()))
            # features.extend(list(Ptm_mutation.ptm_pdbchain_feature.values()))
            # features.extend(list(Ptm_mutation.ptm_uniprot_feature.values()))
            # features.extend(list(Ptm_mutation.mutation_pdbchain_feature.values()))
            # features.extend(list(Ptm_mutation.mutation_uniprot_feature.values()))
            features.extend(Ptm_CrossTalk.features)
            print(Ptm_CrossTalk.features)
            print(features)
            # print(len(features))
            # print(len(Ptm_mutation.features))
            # print(len(titles))
            sheet1.append(features)

    Ptm_CrossTalk_excel.save(featurefile)

# Save feature sets separately
get_feature(positiveset_file, 'PositiveSet_feature.xls')
get_feature(negativeset_file, 'NegativeSet_feature.xls')
